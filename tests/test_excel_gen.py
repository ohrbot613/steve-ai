"""Tests for src/excel_gen.py — verify workbook structure, styling, and content.

These tests re-open generated .xlsx files via openpyxl so we are validating the
real serialized output, not just in-memory attributes. Maps to issue #29
acceptance criteria:
  * 5-tab reconciliation workbook generates correctly
  * conditional formatting applies correctly (row-fill colour by match_status)
  * numbers and dates formatted properly (currency / date number formats)
  * dashboard workbook generates for all suppliers
"""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import pytest

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import load_workbook

from src import excel_gen


# ---- fixtures ---------------------------------------------------------------

@pytest.fixture()
def supplier() -> dict:
    return {"id": 1, "name": "Founding IP", "currency": "USD"}


@pytest.fixture()
def statement_invoices() -> list[dict]:
    return [
        {"id": 101, "invoice_number": "INV-001", "invoice_date": "2026-04-01",
         "amount": 100.0, "currency": "USD"},
        {"id": 102, "invoice_number": "INV-002", "invoice_date": "2026-04-15",
         "amount": 250.0, "currency": "USD"},
        {"id": 103, "invoice_number": "INV-003", "invoice_date": "2026-04-20",
         "amount": 500.0, "currency": "USD"},
        {"id": 104, "invoice_number": "INV-004", "invoice_date": "2026-04-25",
         "amount": 75.0, "currency": "USD"},
    ]


@pytest.fixture()
def xero_invoices() -> list[dict]:
    return [
        {"id": 201, "xero_invoice_id": "x-001", "invoice_number": "INV-001",
         "invoice_date": "2026-04-01", "amount": 100.0, "currency": "USD",
         "status": "AUTHORISED"},
        {"id": 202, "xero_invoice_id": "x-002", "invoice_number": "INV-002",
         "invoice_date": "2026-04-15", "amount": 240.0, "currency": "USD",
         "status": "AUTHORISED"},
        {"id": 203, "xero_invoice_id": "x-003", "invoice_number": "INV-003",
         "invoice_date": "2026-04-20", "amount": 500.0, "currency": "EUR",
         "status": "AUTHORISED"},
        {"id": 204, "xero_invoice_id": "x-005", "invoice_number": "INV-005",
         "invoice_date": "2026-04-30", "amount": 999.0, "currency": "USD",
         "status": "PAID"},
    ]


@pytest.fixture()
def statement(statement_invoices, xero_invoices) -> dict:
    return {
        "id": 1,
        "period_start": "2026-04-01",
        "period_end": "2026-04-30",
        "currency": "USD",
        "statement_total": 925.0,
        "statement_invoices": statement_invoices,
        "xero_invoices": xero_invoices,
    }


@pytest.fixture()
def reconciliation_results() -> dict:
    """Hand-built results covering each match_status colour the spec calls out."""
    return {
        "results": [
            {"statement_invoice_id": 101, "xero_invoice_id": 201,
             "match_status": "MATCHED", "match_method": "exact_number",
             "confidence": 0.99, "amount_difference": 0.0,
             "reasoning": "Exact amount + invoice-number match"},
            {"statement_invoice_id": 102, "xero_invoice_id": 202,
             "match_status": "AMOUNT_MISMATCH", "match_method": "exact_number",
             "confidence": 0.7, "amount_difference": 10.0,
             "reasoning": "Amounts differ by 10.00"},
            {"statement_invoice_id": 103, "xero_invoice_id": 203,
             "match_status": "CURRENCY_MISMATCH", "match_method": "exact_number",
             "confidence": 0.5, "amount_difference": 0.0,
             "reasoning": "Currency mismatch: USD vs EUR"},
            {"statement_invoice_id": None, "xero_invoice_id": 204,
             "match_status": "ALREADY_PAID", "match_method": "none",
             "confidence": 0.9, "amount_difference": -999.0,
             "reasoning": "Xero invoice already paid and not on statement"},
            {"statement_invoice_id": 104, "xero_invoice_id": None,
             "match_status": "MISSING_FROM_XERO", "match_method": "none",
             "confidence": 0.0, "amount_difference": 75.0,
             "reasoning": "No matching Xero invoice found"},
        ],
        "balance": {
            "statement_declared_total": 925.0,
            "statement_sum": 925.0,
            "xero_unpaid_sum": 840.0,
            "matched_sum": 100.0,
            "declared_vs_sum_variance": 0.0,
            "unexplained_variance": 85.0,
        },
        "overall_confidence": "MEDIUM",
        "needs_review": [],
    }


@pytest.fixture()
def payment_tier() -> dict:
    return {
        "name": "aggressive",
        "window_days": 30,
        "cutoff_date": "2026-05-31",
        "invoices": [
            {"id": 101, "invoice_number": "INV-001", "invoice_date": "2026-04-01",
             "due_date": "2026-05-01", "amount": 100.0, "currency": "USD"},
            {"id": 102, "invoice_number": "INV-002", "invoice_date": "2026-04-15",
             "due_date": "2026-05-15", "amount": 250.0, "currency": "USD"},
        ],
        "total": 350.0,
        "invoice_count": 2,
    }


@pytest.fixture()
def audit_entries() -> list[dict]:
    return [
        {"created_at": "2026-05-01 09:00:00", "actor": "statement_parser",
         "action": "parse", "entity_type": "statement", "entity_id": "1",
         "payload": '{"invoice_count": 4}'},
        {"created_at": "2026-05-01 09:01:00", "actor": "matcher",
         "action": "match", "entity_type": "reconciliation", "entity_id": "1",
         "payload": '{"matched": 1, "discrepancies": 3}'},
    ]


@pytest.fixture()
def recon_wb_path(
    tmp_path: Path,
    supplier, statement, reconciliation_results, payment_tier, audit_entries,
) -> Path:
    out = tmp_path / "recon.xlsx"
    return excel_gen.generate_reconciliation_workbook(
        supplier, statement, reconciliation_results, payment_tier, audit_entries,
        output_path=out,
    )


@pytest.fixture()
def recon_wb(recon_wb_path):
    return load_workbook(recon_wb_path)


# ---- structure --------------------------------------------------------------

class TestReconciliationWorkbookStructure:
    def test_file_is_written_to_specified_path(self, recon_wb_path: Path):
        assert recon_wb_path.exists()
        assert recon_wb_path.stat().st_size > 0

    def test_has_exactly_five_named_tabs_in_spec_order(self, recon_wb):
        assert recon_wb.sheetnames == [
            "Summary",
            "Invoice Matching",
            "Payment Schedule",
            "Discrepancies",
            "Audit Log",
        ]

    def test_default_blank_sheet_is_removed(self, recon_wb):
        # openpyxl always creates a "Sheet" — our generator should drop it.
        assert "Sheet" not in recon_wb.sheetnames


# ---- headers / freeze / styling --------------------------------------------

class TestHeaderStyling:
    @pytest.mark.parametrize("tab", [
        "Summary", "Invoice Matching", "Payment Schedule", "Discrepancies", "Audit Log",
    ])
    def test_first_row_is_frozen(self, recon_wb, tab):
        ws = recon_wb[tab]
        assert ws.freeze_panes == "A2"

    @pytest.mark.parametrize("tab", [
        "Summary", "Invoice Matching", "Payment Schedule", "Discrepancies", "Audit Log",
    ])
    def test_header_cells_are_bold(self, recon_wb, tab):
        ws = recon_wb[tab]
        # Read at least the first header cell.
        cell = ws.cell(row=1, column=1)
        assert cell.font.bold is True

    @pytest.mark.parametrize("tab", [
        "Summary", "Invoice Matching", "Payment Schedule", "Discrepancies", "Audit Log",
    ])
    def test_header_cells_have_dark_blue_fill_and_white_text(self, recon_wb, tab):
        ws = recon_wb[tab]
        cell = ws.cell(row=1, column=1)
        fill_color = cell.fill.fgColor.rgb or ""
        # openpyxl stores ARGB ("FF" alpha prefix); accept either form.
        assert fill_color.upper().endswith(excel_gen.HEADER_FILL_HEX)
        font_color = (cell.font.color.rgb or "").upper() if cell.font.color else ""
        assert font_color.endswith(excel_gen.HEADER_FONT_HEX)


class TestColumnWidths:
    def test_column_widths_are_set_on_invoice_matching_tab(self, recon_wb):
        ws = recon_wb["Invoice Matching"]
        # All 13 columns of the Invoice Matching tab should have an explicit width.
        for letter in "ABCDEFGHIJKLM":
            assert ws.column_dimensions[letter].width is not None
            assert ws.column_dimensions[letter].width >= 10  # min clamp from generator


# ---- Summary tab content ----------------------------------------------------

class TestSummaryTab:
    def test_includes_supplier_period_and_totals(self, recon_wb):
        ws = recon_wb["Summary"]
        labels = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert "Supplier" in labels
        assert "Statement period start" in labels
        assert "Statement period end" in labels
        assert "Statement total" in labels
        assert "Xero unpaid total" in labels
        assert "Variance (statement − Xero)" in labels
        assert "Overall confidence" in labels
        assert "Payment amount" in labels
        assert "Reconciliation date" in labels

    def test_statement_total_uses_currency_format(self, recon_wb):
        ws = recon_wb["Summary"]
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "Statement total":
                cell = ws.cell(row=r, column=2)
                assert cell.value == 925.0
                assert "#,##0.00" in cell.number_format
                assert "$" in cell.number_format  # USD symbol
                return
        pytest.fail("Statement total row not found")

    def test_variance_is_statement_total_minus_xero_unpaid(self, recon_wb):
        ws = recon_wb["Summary"]
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "Variance (statement − Xero)":
                assert ws.cell(row=r, column=2).value == pytest.approx(85.0)
                return
        pytest.fail("Variance row not found")

    def test_period_dates_use_date_format(self, recon_wb):
        ws = recon_wb["Summary"]
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "Statement period start":
                cell = ws.cell(row=r, column=2)
                assert cell.number_format == excel_gen.DATE_FORMAT
                # Excel doesn't distinguish date vs datetime — openpyxl rehydrates
                # a date cell as midnight datetime on load.
                assert cell.value == datetime(2026, 4, 1)
                return
        pytest.fail("Statement period start row not found")


# ---- Invoice Matching tab ---------------------------------------------------

class TestInvoiceMatchingTab:
    def test_headers_match_spec(self, recon_wb):
        ws = recon_wb["Invoice Matching"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 14)]
        assert headers == [
            "Invoice #",
            "Statement Date",
            "Statement Amt",
            "Statement CCY",
            "Xero Invoice #",
            "Xero Date",
            "Xero Amt",
            "Xero CCY",
            "Xero Status",
            "Match Status",
            "Confidence %",
            "Difference",
            "Notes",
        ]

    def test_one_data_row_per_reconciliation_result(self, recon_wb, reconciliation_results):
        ws = recon_wb["Invoice Matching"]
        # max_row counts the header.
        assert ws.max_row == 1 + len(reconciliation_results["results"])

    def test_statement_date_cells_are_date_typed_with_format(self, recon_wb):
        ws = recon_wb["Invoice Matching"]
        cell = ws.cell(row=2, column=2)  # first data row, "Statement Date"
        assert cell.number_format == excel_gen.DATE_FORMAT
        assert cell.value == datetime(2026, 4, 1)

    def test_amount_cells_use_currency_format(self, recon_wb):
        ws = recon_wb["Invoice Matching"]
        # First row corresponds to the MATCHED USD pair (statement amt = 100, USD).
        cell = ws.cell(row=2, column=3)
        assert cell.value == 100.0
        assert "#,##0.00" in cell.number_format
        assert "$" in cell.number_format

    def test_confidence_cell_uses_percent_format(self, recon_wb):
        ws = recon_wb["Invoice Matching"]
        cell = ws.cell(row=2, column=11)  # Confidence %
        assert cell.number_format == excel_gen.PERCENT_FORMAT
        assert cell.value == pytest.approx(0.99)


# ---- Conditional row colouring (THE spec acceptance criterion) --------------

class TestConditionalRowColours:
    """Each match_status colours the entire row per the issue spec colour key."""

    @pytest.fixture()
    def status_rows(self, recon_wb):
        ws = recon_wb["Invoice Matching"]
        out: dict = {}
        for r in range(2, ws.max_row + 1):
            status = ws.cell(row=r, column=10).value  # "Match Status" column
            out[status] = r
        return ws, out

    def _row_fill(self, ws, row_idx: int, col_idx: int = 1) -> str:
        return (ws.cell(row=row_idx, column=col_idx).fill.fgColor.rgb or "").upper()

    def test_matched_row_is_green(self, status_rows):
        ws, rows = status_rows
        assert self._row_fill(ws, rows["MATCHED"]).endswith("C6EFCE")

    def test_amount_mismatch_row_is_yellow(self, status_rows):
        ws, rows = status_rows
        assert self._row_fill(ws, rows["AMOUNT_MISMATCH"]).endswith("FFEB9C")

    def test_currency_mismatch_row_is_red(self, status_rows):
        ws, rows = status_rows
        assert self._row_fill(ws, rows["CURRENCY_MISMATCH"]).endswith("FFC7CE")

    def test_already_paid_row_is_blue(self, status_rows):
        ws, rows = status_rows
        assert self._row_fill(ws, rows["ALREADY_PAID"]).endswith("BDD7EE")

    def test_missing_from_xero_row_is_orange(self, status_rows):
        ws, rows = status_rows
        assert self._row_fill(ws, rows["MISSING_FROM_XERO"]).endswith("FFD9B3")

    def test_fill_extends_across_every_column_in_the_row(self, status_rows):
        ws, rows = status_rows
        matched_row = rows["MATCHED"]
        # 13 columns on this tab; every one of them should carry the same fill.
        for c in range(1, 14):
            assert self._row_fill(ws, matched_row, c).endswith("C6EFCE")


# ---- Payment Schedule tab ---------------------------------------------------

class TestPaymentScheduleTab:
    def test_headers(self, recon_wb):
        ws = recon_wb["Payment Schedule"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 9)]
        assert headers == [
            "Invoice #", "Invoice Date", "Due Date", "Days Overdue",
            "Amount", "Currency", "Tier", "Include?",
        ]

    def test_one_row_per_payment_tier_invoice(self, recon_wb, payment_tier):
        ws = recon_wb["Payment Schedule"]
        assert ws.max_row == 1 + len(payment_tier["invoices"])

    def test_due_date_uses_date_format(self, recon_wb):
        ws = recon_wb["Payment Schedule"]
        cell = ws.cell(row=2, column=3)
        assert cell.number_format == excel_gen.DATE_FORMAT
        assert cell.value == datetime(2026, 5, 1)

    def test_tier_name_appears_on_every_row(self, recon_wb):
        ws = recon_wb["Payment Schedule"]
        tier_values = {ws.cell(row=r, column=7).value for r in range(2, ws.max_row + 1)}
        assert tier_values == {"aggressive"}


# ---- Discrepancies tab ------------------------------------------------------

class TestDiscrepanciesTab:
    def test_contains_only_non_matched_results(self, recon_wb, reconciliation_results):
        ws = recon_wb["Discrepancies"]
        non_matched = [r for r in reconciliation_results["results"]
                       if r["match_status"] != "MATCHED"]
        assert ws.max_row == 1 + len(non_matched)
        # "Type" column should be the non-MATCHED statuses only.
        types = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
        assert "MATCHED" not in types
        assert types == {r["match_status"] for r in non_matched}

    def test_action_required_is_populated(self, recon_wb):
        ws = recon_wb["Discrepancies"]
        for r in range(2, ws.max_row + 1):
            assert ws.cell(row=r, column=6).value  # Action Required column

    def test_rows_carry_status_fill_colour(self, recon_wb):
        ws = recon_wb["Discrepancies"]
        # CURRENCY_MISMATCH row must be red.
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "CURRENCY_MISMATCH":
                fill = (ws.cell(row=r, column=1).fill.fgColor.rgb or "").upper()
                assert fill.endswith("FFC7CE")
                return
        pytest.fail("CURRENCY_MISMATCH row not present")


# ---- Audit Log tab ----------------------------------------------------------

class TestAuditLogTab:
    def test_headers(self, recon_wb):
        ws = recon_wb["Audit Log"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 7)]
        assert headers == ["Timestamp", "Actor", "Action", "Entity Type",
                           "Entity ID", "Payload"]

    def test_one_row_per_audit_entry(self, recon_wb, audit_entries):
        ws = recon_wb["Audit Log"]
        assert ws.max_row == 1 + len(audit_entries)

    def test_actor_and_action_are_written(self, recon_wb, audit_entries):
        ws = recon_wb["Audit Log"]
        actors = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        actions = [ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1)]
        assert actors == [e["actor"] for e in audit_entries]
        assert actions == [e["action"] for e in audit_entries]


# ---- Print area -------------------------------------------------------------

class TestPrintArea:
    @pytest.mark.parametrize("tab", [
        "Summary", "Invoice Matching", "Payment Schedule", "Discrepancies", "Audit Log",
    ])
    def test_print_area_is_set(self, recon_wb, tab):
        ws = recon_wb[tab]
        # openpyxl normalizes print_area to a string like "'Tab'!$A$1:$M$6".
        assert ws.print_area is not None and ws.print_area != ""


# ---- Empty / missing data ---------------------------------------------------

class TestEmptyAndMissingData:
    def test_no_payment_tier_still_produces_workbook(
        self, tmp_path, supplier, statement, reconciliation_results, audit_entries,
    ):
        path = excel_gen.generate_reconciliation_workbook(
            supplier, statement, reconciliation_results,
            payment_tier=None, audit_entries=audit_entries,
            output_path=tmp_path / "no_tier.xlsx",
        )
        wb = load_workbook(path)
        ws = wb["Payment Schedule"]
        # Headers present, no data rows.
        assert ws.max_row == 1
        assert ws.cell(row=1, column=1).value == "Invoice #"

    def test_no_audit_entries_still_produces_workbook(
        self, tmp_path, supplier, statement, reconciliation_results, payment_tier,
    ):
        path = excel_gen.generate_reconciliation_workbook(
            supplier, statement, reconciliation_results, payment_tier, audit_entries=[],
            output_path=tmp_path / "no_audit.xlsx",
        )
        wb = load_workbook(path)
        ws = wb["Audit Log"]
        assert ws.max_row == 1
        assert ws.cell(row=1, column=1).value == "Timestamp"

    def test_empty_reconciliation_still_renders_all_five_tabs(
        self, tmp_path, supplier, statement,
    ):
        empty_recon = {"results": [], "balance": {}, "overall_confidence": "LOW",
                       "needs_review": []}
        path = excel_gen.generate_reconciliation_workbook(
            supplier, statement, empty_recon, payment_tier=None, audit_entries=None,
            output_path=tmp_path / "empty.xlsx",
        )
        wb = load_workbook(path)
        assert wb.sheetnames == [
            "Summary", "Invoice Matching", "Payment Schedule",
            "Discrepancies", "Audit Log",
        ]

    def test_default_output_path_uses_slugged_supplier_name(
        self, tmp_path, monkeypatch, supplier, statement, reconciliation_results,
    ):
        monkeypatch.chdir(tmp_path)
        path = excel_gen.generate_reconciliation_workbook(
            supplier, statement, reconciliation_results,
        )
        assert path.parent == tmp_path
        assert path.name.startswith("reconciliation_Founding_IP_")
        assert path.name.endswith(".xlsx")


# ---- Dashboard workbook -----------------------------------------------------

@pytest.fixture()
def dashboard_path(tmp_path: Path) -> Path:
    suppliers = [
        {"supplier_name": "Founding IP", "currency": "USD",
         "last_statement_at": "2026-05-01", "open_discrepancies": 3,
         "outstanding_amount": 5000.0},
        {"supplier_name": "Maersk", "currency": "EUR",
         "last_statement_at": "2026-04-28", "open_discrepancies": 0,
         "outstanding_amount": 0.0},
    ]
    outstanding = [
        {"supplier_name": "Founding IP", "invoice_number": "INV-002",
         "invoice_date": "2026-04-15", "due_date": "2026-05-15",
         "amount": 250.0, "currency": "USD", "status": "AUTHORISED"},
        {"supplier_name": "Founding IP", "invoice_number": "INV-004",
         "invoice_date": "2026-04-25", "due_date": "2026-05-25",
         "amount": 75.0, "currency": "USD", "status": "AUTHORISED"},
    ]
    discrepancies = [
        {"supplier_name": "Founding IP", "match_status": "AMOUNT_MISMATCH",
         "amount_difference": 10.0, "confidence": 0.7,
         "created_at": "2026-05-01", "reasoning": "Amounts differ by 10.00"},
        {"supplier_name": "Founding IP", "match_status": "CURRENCY_MISMATCH",
         "amount_difference": 0.0, "confidence": 0.5,
         "created_at": "2026-05-01", "reasoning": "Currency mismatch: USD vs EUR"},
    ]
    payments = [
        {"created_at": "2026-05-02", "supplier_name": "Founding IP",
         "decision_type": "PAY", "amount": 100.0, "currency": "USD",
         "rationale": "Auto-pay MATCHED tier"},
    ]
    out = tmp_path / "dashboard.xlsx"
    return excel_gen.generate_dashboard_workbook(
        suppliers,
        outstanding_invoices=outstanding,
        discrepancies=discrepancies,
        payment_history=payments,
        output_path=out,
    )


class TestDashboardWorkbook:
    def test_has_four_tabs_in_expected_order(self, dashboard_path):
        wb = load_workbook(dashboard_path)
        assert wb.sheetnames == [
            "All Suppliers",
            "Outstanding Invoices",
            "Open Discrepancies",
            "Payment History",
        ]

    def test_all_suppliers_tab_has_one_row_per_supplier(self, dashboard_path):
        wb = load_workbook(dashboard_path)
        ws = wb["All Suppliers"]
        assert ws.max_row == 3  # header + 2 suppliers
        assert ws.cell(row=2, column=1).value == "Founding IP"
        assert ws.cell(row=3, column=1).value == "Maersk"

    def test_outstanding_amount_uses_currency_format(self, dashboard_path):
        wb = load_workbook(dashboard_path)
        ws = wb["All Suppliers"]
        cell = ws.cell(row=2, column=5)  # USD supplier outstanding amount
        assert cell.value == 5000.0
        assert "$" in cell.number_format

    def test_outstanding_invoices_tab_has_data(self, dashboard_path):
        wb = load_workbook(dashboard_path)
        ws = wb["Outstanding Invoices"]
        assert ws.max_row == 3  # header + 2 invoices
        # Due date column uses the date format.
        assert ws.cell(row=2, column=4).number_format == excel_gen.DATE_FORMAT

    def test_discrepancies_tab_has_status_fill(self, dashboard_path):
        wb = load_workbook(dashboard_path)
        ws = wb["Open Discrepancies"]
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=2).value == "CURRENCY_MISMATCH":
                fill = (ws.cell(row=r, column=1).fill.fgColor.rgb or "").upper()
                assert fill.endswith("FFC7CE")
                return
        pytest.fail("CURRENCY_MISMATCH row missing")

    def test_payment_history_tab_includes_decision_rows(self, dashboard_path):
        wb = load_workbook(dashboard_path)
        ws = wb["Payment History"]
        assert ws.max_row == 2
        assert ws.cell(row=2, column=3).value == "PAY"

    def test_empty_suppliers_still_renders_four_tabs(self, tmp_path):
        path = excel_gen.generate_dashboard_workbook(
            [], output_path=tmp_path / "empty_dashboard.xlsx",
        )
        wb = load_workbook(path)
        assert wb.sheetnames == [
            "All Suppliers",
            "Outstanding Invoices",
            "Open Discrepancies",
            "Payment History",
        ]
        # Just headers on each tab.
        for tab in wb.sheetnames:
            assert wb[tab].max_row == 1


# ---- openpyxl-missing branch ------------------------------------------------

class TestOpenpyxlMissing:
    def test_clear_error_when_openpyxl_unavailable(self, monkeypatch, tmp_path,
                                                   supplier, statement,
                                                   reconciliation_results):
        def _broken_loader():
            raise ImportError(excel_gen._OPENPYXL_ERROR)

        monkeypatch.setattr(excel_gen, "_load_openpyxl", _broken_loader)
        with pytest.raises(ImportError, match="openpyxl is required"):
            excel_gen.generate_reconciliation_workbook(
                supplier, statement, reconciliation_results,
                output_path=tmp_path / "nope.xlsx",
            )
