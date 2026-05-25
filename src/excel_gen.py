"""Excel workbook generator — Jeffrey's primary deliverable.

Produces two workbook flavours::

    generate_reconciliation_workbook(supplier, statement, reconciliation_results,
                                     payment_tier, audit_entries) -> Path

Five tabs (Summary, Invoice Matching, Payment Schedule, Discrepancies, Audit Log)
with conditional row fills by match_status, frozen + styled headers, currency /
date number formats, and auto-fitted column widths.

    generate_dashboard_workbook(all_suppliers_status, *, outstanding_invoices=...,
                                discrepancies=..., payment_history=...) -> Path

Four tabs giving a CFO-wide view across every supplier.

openpyxl is loaded lazily so the rest of the codebase keeps importing cleanly in
environments that don't need Excel output (CI, tests for other modules). Calling
either generator without openpyxl raises a clear ImportError pointing at the
install command.
"""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Optional, Sequence

# ---- match-status colour map (issue #29) ------------------------------------
STATUS_FILL_COLORS = {
    "MATCHED": "C6EFCE",            # green
    "AMOUNT_MISMATCH": "FFEB9C",    # yellow
    "CURRENCY_MISMATCH": "FFC7CE",  # red
    "ALREADY_PAID": "BDD7EE",       # blue
    "MISSING_FROM_XERO": "FFD9B3",  # orange
    "MISSING_FROM_STATEMENT": "D9D9D9",  # grey
    "AMBIGUOUS": "E4DFEC",          # lavender (not in spec, but matcher emits it)
}

# ---- format strings ----------------------------------------------------------
CURRENCY_FORMAT = '#,##0.00'
CURRENCY_FORMAT_WITH_SYMBOL = {
    "USD": '"$"#,##0.00',
    "GBP": '"£"#,##0.00',
    "EUR": '"€"#,##0.00',
    "EGP": '"E£ "#,##0.00',
}
DATE_FORMAT = 'DD-MMM-YYYY'
PERCENT_FORMAT = '0.00%'

HEADER_FILL_HEX = "1F4E78"   # dark blue
HEADER_FONT_HEX = "FFFFFF"   # white


# ---- openpyxl loader ---------------------------------------------------------

_OPENPYXL_ERROR = (
    "openpyxl is required to generate Excel workbooks. "
    "Install with: pip install openpyxl"
)


def _load_openpyxl():
    """Import openpyxl lazily so other modules don't pay the import cost."""
    try:
        import openpyxl  # noqa: F401
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - depends on environment
        raise ImportError(_OPENPYXL_ERROR) from exc
    return {
        "Workbook": Workbook,
        "Alignment": Alignment,
        "Font": Font,
        "PatternFill": PatternFill,
        "get_column_letter": get_column_letter,
    }


# ---- value coercion ---------------------------------------------------------

def _as_date(value) -> Optional[date]:
    """Coerce common date shapes to a datetime.date; None passes through."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        # ISO is by far the most common shape downstream — keep this cheap.
        try:
            return datetime.fromisoformat(value).date()
        except ValueError:
            for fmt in ("%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y", "%m/%d/%Y"):
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
    return None


def _as_float(value) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _currency_format(currency: Optional[str]) -> str:
    if not currency:
        return CURRENCY_FORMAT
    return CURRENCY_FORMAT_WITH_SYMBOL.get(currency.upper(), CURRENCY_FORMAT)


# ---- shared styling ---------------------------------------------------------

def _style_header_row(ws, headers: Sequence[str], deps: dict) -> None:
    """Write headers in row 1, then style + freeze them."""
    Font = deps["Font"]
    PatternFill = deps["PatternFill"]
    Alignment = deps["Alignment"]
    header_font = Font(bold=True, color=HEADER_FONT_HEX)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_HEX)
    header_align = Alignment(horizontal="left", vertical="center")
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws.freeze_panes = "A2"


def _autofit_columns(ws, headers: Sequence[str], deps: dict) -> None:
    """Set per-column widths from the longest cell value in each column.

    Cheap approximation — openpyxl can't actually auto-fit, so we measure
    rendered string length per cell and clamp the result. Good enough for a
    CFO scanning the workbook in Excel/Sheets.
    """
    get_column_letter = deps["get_column_letter"]
    for col_idx, _ in enumerate(headers, start=1):
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            value = row[0]
            if value is None:
                continue
            if isinstance(value, (date, datetime)):
                rendered = value.strftime("%d-%b-%Y")
            elif isinstance(value, float):
                rendered = f"{value:,.2f}"
            else:
                rendered = str(value)
            if len(rendered) > max_len:
                max_len = len(rendered)
        # +2 for padding; clamp to avoid runaway widths on freeform "reasoning".
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)


def _apply_status_fill(ws, row_idx: int, col_count: int, status: Optional[str], deps: dict) -> None:
    if not status:
        return
    hex_color = STATUS_FILL_COLORS.get(status.upper())
    if not hex_color:
        return
    fill = deps["PatternFill"]("solid", fgColor=hex_color)
    for col_idx in range(1, col_count + 1):
        ws.cell(row=row_idx, column=col_idx).fill = fill


def _set_print_area(ws, headers: Sequence[str], row_count: int, deps: dict) -> None:
    if row_count <= 0:
        return
    last_col = deps["get_column_letter"](len(headers))
    ws.print_area = f"A1:{last_col}{row_count + 1}"  # +1 for header row


# ---- tab builders (reconciliation workbook) ---------------------------------

def _build_summary_tab(
    ws,
    *,
    supplier: dict,
    statement: dict,
    reconciliation_results: dict,
    payment_tier: Optional[dict],
    deps: dict,
) -> None:
    headers = ["Field", "Value"]
    _style_header_row(ws, headers, deps)

    balance = reconciliation_results.get("balance", {}) if reconciliation_results else {}
    overall_conf = reconciliation_results.get("overall_confidence") if reconciliation_results else None

    statement_total = _as_float(
        statement.get("statement_total")
        or balance.get("statement_declared_total")
        or balance.get("statement_sum")
    )
    xero_total = _as_float(balance.get("xero_unpaid_sum"))
    variance = None
    if statement_total is not None and xero_total is not None:
        variance = round(statement_total - xero_total, 2)
    elif balance.get("unexplained_variance") is not None:
        variance = _as_float(balance.get("unexplained_variance"))

    currency = (statement.get("currency") or supplier.get("currency") or "").upper() or None
    money_fmt = _currency_format(currency)

    payment_amount = None
    payment_tier_name = None
    if payment_tier:
        payment_amount = _as_float(payment_tier.get("total"))
        payment_tier_name = payment_tier.get("name")

    rows: list[tuple[str, Any, Optional[str]]] = [
        ("Supplier", supplier.get("name"), None),
        ("Supplier currency", currency, None),
        ("Statement period start", _as_date(statement.get("period_start")), DATE_FORMAT),
        ("Statement period end", _as_date(statement.get("period_end")), DATE_FORMAT),
        ("Statement total", statement_total, money_fmt),
        ("Xero unpaid total", xero_total, money_fmt),
        ("Variance (statement − Xero)", variance, money_fmt),
        ("Overall confidence", overall_conf, None),
        ("Payment tier", payment_tier_name, None),
        ("Payment amount", payment_amount, money_fmt),
        ("Reconciliation date", date.today(), DATE_FORMAT),
    ]

    for row_idx, (label, value, number_format) in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=label)
        cell = ws.cell(row=row_idx, column=2, value=value)
        if number_format:
            cell.number_format = number_format

    _autofit_columns(ws, headers, deps)
    _set_print_area(ws, headers, len(rows), deps)


def _build_invoice_matching_tab(
    ws,
    *,
    reconciliation_results: dict,
    statement_invoices_by_id: dict,
    xero_invoices_by_id: dict,
    statement_currency: Optional[str],
    deps: dict,
) -> None:
    headers = [
        "Invoice #",
        "Statement Date",
        "Statement Amt",
        "Statement CCY",
        "Xero Invoice #",
        "Xero Date",
        "Xero Amt",
        "Xero CCY",
        "Xero Status",
        "Match Status",
        "Confidence %",
        "Difference",
        "Notes",
    ]
    _style_header_row(ws, headers, deps)

    money_fmt = _currency_format(statement_currency)

    results = (reconciliation_results or {}).get("results", []) if reconciliation_results else []

    for row_offset, row_data in enumerate(results, start=2):
        s = statement_invoices_by_id.get(row_data.get("statement_invoice_id"), {}) or {}
        x = xero_invoices_by_id.get(row_data.get("xero_invoice_id"), {}) or {}

        cells: list[tuple[Any, Optional[str]]] = [
            (s.get("invoice_number") or x.get("invoice_number"), None),
            (_as_date(s.get("invoice_date")), DATE_FORMAT),
            (_as_float(s.get("amount")), _currency_format(s.get("currency") or statement_currency)),
            ((s.get("currency") or "").upper() or None, None),
            (x.get("invoice_number"), None),
            (_as_date(x.get("invoice_date")), DATE_FORMAT),
            (_as_float(x.get("amount")), _currency_format(x.get("currency") or statement_currency)),
            ((x.get("currency") or "").upper() or None, None),
            (x.get("status"), None),
            (row_data.get("match_status"), None),
            (_as_float(row_data.get("confidence")), PERCENT_FORMAT),
            (_as_float(row_data.get("amount_difference")), money_fmt),
            (row_data.get("reasoning"), None),
        ]
        for col_idx, (value, number_format) in enumerate(cells, start=1):
            cell = ws.cell(row=row_offset, column=col_idx, value=value)
            if number_format:
                cell.number_format = number_format

        _apply_status_fill(ws, row_offset, len(headers), row_data.get("match_status"), deps)

    _autofit_columns(ws, headers, deps)
    _set_print_area(ws, headers, len(results), deps)


def _build_payment_schedule_tab(
    ws,
    *,
    payment_tier: Optional[dict],
    deps: dict,
) -> None:
    headers = [
        "Invoice #",
        "Invoice Date",
        "Due Date",
        "Days Overdue",
        "Amount",
        "Currency",
        "Tier",
        "Include?",
    ]
    _style_header_row(ws, headers, deps)

    invoices = (payment_tier or {}).get("invoices", []) if payment_tier else []
    tier_name = (payment_tier or {}).get("name", "")
    today = date.today()

    for row_offset, inv in enumerate(invoices, start=2):
        due = _as_date(inv.get("due_date"))
        days_overdue = (today - due).days if due else None
        if days_overdue is not None and days_overdue < 0:
            days_overdue = 0

        cells: list[tuple[Any, Optional[str]]] = [
            (inv.get("invoice_number"), None),
            (_as_date(inv.get("invoice_date")), DATE_FORMAT),
            (due, DATE_FORMAT),
            (days_overdue, None),
            (_as_float(inv.get("amount")), _currency_format(inv.get("currency"))),
            ((inv.get("currency") or "").upper() or None, None),
            (tier_name, None),
            ("INCLUDE", None),
        ]
        for col_idx, (value, number_format) in enumerate(cells, start=1):
            cell = ws.cell(row=row_offset, column=col_idx, value=value)
            if number_format:
                cell.number_format = number_format

    _autofit_columns(ws, headers, deps)
    _set_print_area(ws, headers, len(invoices), deps)


def _build_discrepancies_tab(
    ws,
    *,
    reconciliation_results: dict,
    statement_invoices_by_id: dict,
    xero_invoices_by_id: dict,
    statement_currency: Optional[str],
    deps: dict,
) -> None:
    headers = [
        "Type",
        "Invoice #",
        "Statement Amt",
        "Xero Amt",
        "Difference",
        "Action Required",
        "Status",
        "Notes",
    ]
    _style_header_row(ws, headers, deps)

    results = (reconciliation_results or {}).get("results", []) if reconciliation_results else []
    discrepancies = [r for r in results if r.get("match_status") != "MATCHED"]

    action_for_status = {
        "AMOUNT_MISMATCH": "Reconcile amount difference",
        "CURRENCY_MISMATCH": "Verify currency before paying",
        "ALREADY_PAID": "Confirm payment, no action",
        "MISSING_FROM_XERO": "Add invoice to Xero",
        "MISSING_FROM_STATEMENT": "Request from supplier",
        "AMBIGUOUS": "Manual review",
    }

    money_fmt = _currency_format(statement_currency)

    for row_offset, row_data in enumerate(discrepancies, start=2):
        s = statement_invoices_by_id.get(row_data.get("statement_invoice_id"), {}) or {}
        x = xero_invoices_by_id.get(row_data.get("xero_invoice_id"), {}) or {}
        status = row_data.get("match_status", "")

        cells: list[tuple[Any, Optional[str]]] = [
            (status, None),
            (s.get("invoice_number") or x.get("invoice_number"), None),
            (_as_float(s.get("amount")), money_fmt),
            (_as_float(x.get("amount")), money_fmt),
            (_as_float(row_data.get("amount_difference")), money_fmt),
            (action_for_status.get(status, "Review"), None),
            ("OPEN", None),
            (row_data.get("reasoning"), None),
        ]
        for col_idx, (value, number_format) in enumerate(cells, start=1):
            cell = ws.cell(row=row_offset, column=col_idx, value=value)
            if number_format:
                cell.number_format = number_format

        _apply_status_fill(ws, row_offset, len(headers), status, deps)

    _autofit_columns(ws, headers, deps)
    _set_print_area(ws, headers, len(discrepancies), deps)


def _build_audit_log_tab(
    ws,
    *,
    audit_entries: Sequence[dict],
    deps: dict,
) -> None:
    headers = ["Timestamp", "Actor", "Action", "Entity Type", "Entity ID", "Payload"]
    _style_header_row(ws, headers, deps)

    entries = list(audit_entries or [])

    for row_offset, entry in enumerate(entries, start=2):
        created_at = entry.get("created_at") or entry.get("timestamp")
        # Audit timestamps from sqlite are "YYYY-MM-DD HH:MM:SS" — keep them as strings
        # so the Excel cell renders the full timestamp without timezone trickery.
        cells: list[tuple[Any, Optional[str]]] = [
            (created_at, None),
            (entry.get("actor"), None),
            (entry.get("action"), None),
            (entry.get("entity_type"), None),
            (entry.get("entity_id"), None),
            (entry.get("payload"), None),
        ]
        for col_idx, (value, number_format) in enumerate(cells, start=1):
            cell = ws.cell(row=row_offset, column=col_idx, value=value)
            if number_format:
                cell.number_format = number_format

    _autofit_columns(ws, headers, deps)
    _set_print_area(ws, headers, len(entries), deps)


# ---- index helpers ----------------------------------------------------------

def _index_by_id(invoices: Iterable[dict]) -> dict:
    """Map invoice id -> invoice dict so we can join matcher rows back to fields."""
    out: dict = {}
    for inv in invoices or []:
        key = inv.get("id") or inv.get("statement_invoice_id") or inv.get("xero_invoice_id")
        if key is not None:
            out[key] = inv
    return out


# ---- public API: reconciliation workbook ------------------------------------

def generate_reconciliation_workbook(
    supplier: dict,
    statement: dict,
    reconciliation_results: dict,
    payment_tier: Optional[dict] = None,
    audit_entries: Optional[Sequence[dict]] = None,
    *,
    output_path: Optional[str | Path] = None,
) -> Path:
    """Build the five-tab reconciliation workbook for one supplier statement.

    ``supplier``     — dict with at least ``name`` and ``currency``.
    ``statement``    — dict with ``period_start``, ``period_end``, ``currency``,
                       ``statement_total``, plus the resolved ``statement_invoices``
                       (list of dicts with ``id``, ``invoice_number``,
                       ``invoice_date``, ``amount``, ``currency``) and
                       ``xero_invoices`` (same shape plus ``status``).
    ``reconciliation_results`` — output of ``matcher.match_invoices``.
    ``payment_tier`` — one tier dict from ``payment.calculate_payment_tiers``;
                       optional, drives the Payment Schedule tab.
    ``audit_entries`` — list of audit_log rows for this reconciliation.
    ``output_path``  — where to write the .xlsx; defaults to a slugged filename
                       in CWD using supplier name + today's date.
    """
    deps = _load_openpyxl()
    wb = deps["Workbook"]()
    wb.remove(wb.active)  # ditch the default blank sheet

    statement_invoices_by_id = _index_by_id(statement.get("statement_invoices") or [])
    xero_invoices_by_id = _index_by_id(statement.get("xero_invoices") or [])
    statement_currency = (statement.get("currency") or supplier.get("currency") or "USD")

    _build_summary_tab(
        wb.create_sheet("Summary"),
        supplier=supplier,
        statement=statement,
        reconciliation_results=reconciliation_results,
        payment_tier=payment_tier,
        deps=deps,
    )
    _build_invoice_matching_tab(
        wb.create_sheet("Invoice Matching"),
        reconciliation_results=reconciliation_results,
        statement_invoices_by_id=statement_invoices_by_id,
        xero_invoices_by_id=xero_invoices_by_id,
        statement_currency=statement_currency,
        deps=deps,
    )
    _build_payment_schedule_tab(
        wb.create_sheet("Payment Schedule"),
        payment_tier=payment_tier,
        deps=deps,
    )
    _build_discrepancies_tab(
        wb.create_sheet("Discrepancies"),
        reconciliation_results=reconciliation_results,
        statement_invoices_by_id=statement_invoices_by_id,
        xero_invoices_by_id=xero_invoices_by_id,
        statement_currency=statement_currency,
        deps=deps,
    )
    _build_audit_log_tab(
        wb.create_sheet("Audit Log"),
        audit_entries=audit_entries or [],
        deps=deps,
    )

    output = _resolve_output_path(output_path, supplier=supplier)
    wb.save(str(output))
    return output


def _resolve_output_path(output_path, *, supplier: dict, prefix: str = "reconciliation") -> Path:
    if output_path is not None:
        return Path(output_path)
    name = supplier.get("name") or "supplier"
    slug = "".join(ch if ch.isalnum() else "_" for ch in name).strip("_") or "supplier"
    stamp = date.today().strftime("%Y-%m-%d")
    return Path.cwd() / f"{prefix}_{slug}_{stamp}.xlsx"


# ---- public API: dashboard workbook -----------------------------------------

def _build_dashboard_suppliers_tab(ws, suppliers: Sequence[dict], deps: dict) -> None:
    headers = [
        "Supplier",
        "Currency",
        "Last Statement",
        "Open Discrepancies",
        "Outstanding Amount",
    ]
    _style_header_row(ws, headers, deps)

    for row_offset, s in enumerate(suppliers, start=2):
        currency = (s.get("currency") or "").upper() or None
        cells: list[tuple[Any, Optional[str]]] = [
            (s.get("supplier_name") or s.get("name"), None),
            (currency, None),
            (s.get("last_statement_at"), None),
            (s.get("open_discrepancies"), None),
            (_as_float(s.get("outstanding_amount")), _currency_format(currency)),
        ]
        for col_idx, (value, number_format) in enumerate(cells, start=1):
            cell = ws.cell(row=row_offset, column=col_idx, value=value)
            if number_format:
                cell.number_format = number_format

    _autofit_columns(ws, headers, deps)
    _set_print_area(ws, headers, len(suppliers), deps)


def _build_dashboard_outstanding_tab(ws, invoices: Sequence[dict], deps: dict) -> None:
    headers = [
        "Supplier",
        "Invoice #",
        "Invoice Date",
        "Due Date",
        "Amount",
        "Currency",
        "Status",
    ]
    _style_header_row(ws, headers, deps)

    for row_offset, inv in enumerate(invoices, start=2):
        currency = (inv.get("currency") or "").upper() or None
        cells: list[tuple[Any, Optional[str]]] = [
            (inv.get("supplier_name") or inv.get("supplier"), None),
            (inv.get("invoice_number"), None),
            (_as_date(inv.get("invoice_date")), DATE_FORMAT),
            (_as_date(inv.get("due_date")), DATE_FORMAT),
            (_as_float(inv.get("amount")), _currency_format(currency)),
            (currency, None),
            (inv.get("status"), None),
        ]
        for col_idx, (value, number_format) in enumerate(cells, start=1):
            cell = ws.cell(row=row_offset, column=col_idx, value=value)
            if number_format:
                cell.number_format = number_format

    _autofit_columns(ws, headers, deps)
    _set_print_area(ws, headers, len(invoices), deps)


def _build_dashboard_discrepancies_tab(ws, discrepancies: Sequence[dict], deps: dict) -> None:
    headers = [
        "Supplier",
        "Match Status",
        "Difference",
        "Confidence",
        "Created",
        "Reasoning",
    ]
    _style_header_row(ws, headers, deps)

    for row_offset, row in enumerate(discrepancies, start=2):
        status = row.get("match_status")
        cells: list[tuple[Any, Optional[str]]] = [
            (row.get("supplier_name") or row.get("supplier"), None),
            (status, None),
            (_as_float(row.get("amount_difference")), CURRENCY_FORMAT),
            (_as_float(row.get("confidence")), PERCENT_FORMAT),
            (row.get("created_at"), None),
            (row.get("reasoning"), None),
        ]
        for col_idx, (value, number_format) in enumerate(cells, start=1):
            cell = ws.cell(row=row_offset, column=col_idx, value=value)
            if number_format:
                cell.number_format = number_format
        _apply_status_fill(ws, row_offset, len(headers), status, deps)

    _autofit_columns(ws, headers, deps)
    _set_print_area(ws, headers, len(discrepancies), deps)


def _build_dashboard_payment_history_tab(ws, payments: Sequence[dict], deps: dict) -> None:
    headers = [
        "Date",
        "Supplier",
        "Decision Type",
        "Amount",
        "Currency",
        "Rationale",
    ]
    _style_header_row(ws, headers, deps)

    for row_offset, payment in enumerate(payments, start=2):
        currency = (payment.get("currency") or "").upper() or None
        cells: list[tuple[Any, Optional[str]]] = [
            (_as_date(payment.get("created_at") or payment.get("date")), DATE_FORMAT),
            (payment.get("supplier_name") or payment.get("supplier"), None),
            (payment.get("decision_type"), None),
            (_as_float(payment.get("amount")), _currency_format(currency)),
            (currency, None),
            (payment.get("rationale"), None),
        ]
        for col_idx, (value, number_format) in enumerate(cells, start=1):
            cell = ws.cell(row=row_offset, column=col_idx, value=value)
            if number_format:
                cell.number_format = number_format

    _autofit_columns(ws, headers, deps)
    _set_print_area(ws, headers, len(payments), deps)


def generate_dashboard_workbook(
    all_suppliers_status: Sequence[dict],
    *,
    outstanding_invoices: Optional[Sequence[dict]] = None,
    discrepancies: Optional[Sequence[dict]] = None,
    payment_history: Optional[Sequence[dict]] = None,
    output_path: Optional[str | Path] = None,
) -> Path:
    """Build the four-tab CFO dashboard workbook covering every supplier.

    Each list argument is rendered to its own tab. The CLI / orchestrator is
    expected to source these from the relevant `db_ops` views — keeping the
    generator pure makes it trivial to test and to reuse for ad-hoc reports.
    """
    deps = _load_openpyxl()
    wb = deps["Workbook"]()
    wb.remove(wb.active)

    _build_dashboard_suppliers_tab(
        wb.create_sheet("All Suppliers"),
        all_suppliers_status or [],
        deps,
    )
    _build_dashboard_outstanding_tab(
        wb.create_sheet("Outstanding Invoices"),
        outstanding_invoices or [],
        deps,
    )
    _build_dashboard_discrepancies_tab(
        wb.create_sheet("Open Discrepancies"),
        discrepancies or [],
        deps,
    )
    _build_dashboard_payment_history_tab(
        wb.create_sheet("Payment History"),
        payment_history or [],
        deps,
    )

    output = _resolve_output_path(
        output_path,
        supplier={"name": "dashboard"},
        prefix="dashboard",
    )
    wb.save(str(output))
    return output


__all__ = [
    "CURRENCY_FORMAT",
    "CURRENCY_FORMAT_WITH_SYMBOL",
    "DATE_FORMAT",
    "STATUS_FILL_COLORS",
    "generate_dashboard_workbook",
    "generate_reconciliation_workbook",
]
