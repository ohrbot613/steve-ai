"""Statement parser: extract structured invoice data from PDF / Excel / CSV / text.

Optional dependencies (``pdfplumber``, ``openpyxl``) are loaded lazily so the
module imports cleanly in environments that only need the text/CSV path
(notably the test suite). All extraction work returns plain dicts shaped for
:func:`src.db_ops.add_statement_invoices`.

A ``claude_hook`` seam is exposed on :func:`identify_supplier` and
:func:`extract_invoices` so callers can plug in a Claude-powered fallback
without this module taking a hard dependency on the SDK.
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Iterable, Optional, Sequence

from .normalizer import (
    normalize_invoice_number,
    normalize_supplier_name,
    parse_amount,
    parse_date,
)

# ---- public types ------------------------------------------------------------

ParsedRow = dict
ClaudeHook = Callable[[str, dict], Optional[dict]]
"""Signature: hook(stage, context) -> dict | None.

``stage`` is "identify_supplier" or "extract_invoices". ``context`` carries
whatever the caller wants to forward (text excerpt, filename, known aliases).
Hook returns a parsed result dict, or ``None`` to defer to deterministic logic.
"""


@dataclass
class ParsedFile:
    """Raw output of file-type-specific parsing."""
    text: str = ""
    tables: list[list[list[str]]] = field(default_factory=list)
    source: str = ""  # "pdf" | "excel" | "csv" | "text"
    path: Optional[str] = None


@dataclass
class ParsedStatement:
    """Final structured output of the parser."""
    supplier_name_detected: Optional[str]
    supplier_id: Optional[int]
    supplier_confidence: float
    invoices: list[ParsedRow]
    statement_total: Optional[float]
    statement_period: Optional[str]
    currency: Optional[str]
    invoice_count: int
    source: str

    def to_dict(self) -> dict:
        return {
            "supplier_name_detected": self.supplier_name_detected,
            "supplier_id": self.supplier_id,
            "confidence": round(self.supplier_confidence, 3),
            "invoices": self.invoices,
            "invoice_count": self.invoice_count,
            "statement_total": self.statement_total,
            "statement_period": self.statement_period,
            "currency": self.currency,
            "source": self.source,
        }


# ---- regexes -----------------------------------------------------------------

# Invoice numbers: a token that looks like INV-2026-001, FND/2026/042, 2026001,
# REF#42, etc. Letters/digits with optional internal separators, length >= 3,
# and contains at least one digit.
_INVOICE_TOKEN_RE = re.compile(
    r"\b(?:(?:INV|REF|BILL|FAC|FND|INVOICE|NO\.?|#)[\s\-_/:#.]*)?"
    r"([A-Z]{0,6}[\s\-_/.#:]?\d[A-Z0-9\-_/.#:]{1,30})",
    re.IGNORECASE,
)

# Amounts: optional currency symbol/code, digits w/ thousands separators,
# optional decimal, optional negative or parens.
_AMOUNT_RE = re.compile(
    r"""
    (?P<paren>\()?
    (?:(?P<sym>[$£€])\s*|(?P<code>USD|GBP|EUR|EGP)\s*)?
    (?P<num>-?\d{1,3}(?:,\d{3})+(?:\.\d{1,2})?|-?\d+\.\d{1,2}|-?\d+)
    (?(paren)\))
    """,
    re.IGNORECASE | re.VERBOSE,
)

# Dates the normalizer will parse — we just locate plausible substrings.
_DATE_RE = re.compile(
    r"\b("
    r"\d{4}[-/]\d{1,2}[-/]\d{1,2}"            # 2026-01-15
    r"|\d{1,2}[-/]\d{1,2}[-/]\d{2,4}"         # 15/01/2026
    r"|\d{1,2}\s+[A-Za-z]{3,9}\s+\d{2,4}"     # 15 Jan 2026
    r"|[A-Za-z]{3,9}\s+\d{1,2},?\s+\d{2,4}"   # Jan 15, 2026
    r")\b"
)

_CURRENCY_SYMBOL = {"$": "USD", "£": "GBP", "€": "EUR"}

_TOTAL_LINE_RE = re.compile(
    r"(?:statement\s+total|balance\s+due|total\s+due|grand\s+total|amount\s+due|total)"
    r"\s*[:\-]?\s*(.+)",
    re.IGNORECASE,
)
_PERIOD_RE = re.compile(
    r"(?:statement\s+period|period|for(?:\s+the)?\s+(?:month\s+of)?)"
    r"\s*[:\-]?\s*([A-Za-z0-9 ,\-/]+)",
    re.IGNORECASE,
)


# ---- file dispatch -----------------------------------------------------------

def parse_file(path: str | Path) -> ParsedFile:
    """Detect the file type by extension and parse it."""
    p = Path(path)
    suffix = p.suffix.lower()
    if suffix == ".pdf":
        return parse_pdf(p)
    if suffix in (".xlsx", ".xlsm"):
        return parse_excel(p)
    if suffix == ".csv":
        return parse_csv(p)
    return parse_text(p)


def parse_pdf(path: str | Path) -> ParsedFile:
    """Extract text + tables from a PDF using pdfplumber if available.

    Raises :class:`RuntimeError` if pdfplumber is not installed — callers in
    non-PDF environments should route to :func:`parse_text` instead.
    """
    try:
        import pdfplumber  # type: ignore
    except ImportError as exc:  # pragma: no cover - exercised via env without dep
        raise RuntimeError(
            "pdfplumber is required for PDF parsing. Install with: pip install pdfplumber"
        ) from exc

    text_parts: list[str] = []
    tables: list[list[list[str]]] = []
    with pdfplumber.open(str(path)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            if page_text:
                text_parts.append(page_text)
            for tbl in page.extract_tables() or []:
                tables.append([[(cell or "") for cell in row] for row in tbl])
    return ParsedFile(text="\n".join(text_parts), tables=tables, source="pdf", path=str(path))


def parse_excel(path: str | Path) -> ParsedFile:
    """Extract rows from an .xlsx workbook using openpyxl if available."""
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "openpyxl is required for Excel parsing. Install with: pip install openpyxl"
        ) from exc

    wb = load_workbook(str(path), data_only=True, read_only=True)
    tables: list[list[list[str]]] = []
    text_lines: list[str] = []
    for ws in wb.worksheets:
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells = ["" if c is None else str(c) for c in row]
            rows.append(cells)
            text_lines.append("\t".join(cells))
        if rows:
            tables.append(rows)
    return ParsedFile(text="\n".join(text_lines), tables=tables, source="excel", path=str(path))


def parse_csv(path: str | Path) -> ParsedFile:
    text = Path(path).read_text(encoding="utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = [list(r) for r in reader]
    return ParsedFile(text=text, tables=[rows] if rows else [], source="csv", path=str(path))


def parse_text(path: str | Path) -> ParsedFile:
    text = Path(path).read_text(encoding="utf-8", errors="replace")
    return ParsedFile(text=text, tables=[], source="text", path=str(path))


# ---- supplier identification -------------------------------------------------

def _haystack(value: str) -> str:
    """Normalize for substring matching: collapse separators to single spaces."""
    s = normalize_supplier_name(value)
    return re.sub(r"[_\-./]+", " ", s)


def identify_supplier(
    text: str,
    filename: Optional[str] = None,
    *,
    known_aliases: Optional[Iterable[dict]] = None,
    claude_hook: Optional[ClaudeHook] = None,
) -> dict:
    """Identify the supplier from text + filename against a list of known aliases.

    ``known_aliases`` is an iterable of ``{"supplier_id", "supplier_name",
    "alias", "normalized"}`` entries — typically the union of
    ``suppliers.canonical_name`` and ``supplier_aliases.normalized``.

    Returns ``{"supplier_id", "supplier_name", "confidence", "method"}``.
    ``confidence`` is in [0, 1]; ``supplier_id`` is ``None`` when unknown.
    """
    aliases = list(known_aliases or [])
    haystack_filename = _haystack(filename or "")
    excerpt = text[:2000] if text else ""
    haystack_text = _haystack(excerpt)

    best: Optional[dict] = None
    best_conf = 0.0
    best_method = "none"

    for entry in aliases:
        needle = entry.get("normalized") or normalize_supplier_name(entry.get("alias", ""))
        if not needle:
            continue
        # Exact-ish filename hit first (most reliable).
        if needle and needle in haystack_filename:
            conf = 0.95
            method = "filename"
        elif needle and needle in haystack_text:
            # Longer aliases are more discriminating.
            conf = min(0.92, 0.70 + 0.02 * len(needle.split()))
            method = "text"
        else:
            continue
        if conf > best_conf:
            best = entry
            best_conf = conf
            best_method = method

    if best is not None:
        return {
            "supplier_id": best.get("supplier_id"),
            "supplier_name": best.get("supplier_name") or best.get("alias"),
            "confidence": best_conf,
            "method": best_method,
        }

    # Optional Claude fallback — caller owns the network dependency.
    if claude_hook is not None:
        hook_out = claude_hook(
            "identify_supplier",
            {"text": excerpt, "filename": filename, "known_aliases": aliases},
        )
        if hook_out and hook_out.get("supplier_id"):
            return {
                "supplier_id": hook_out["supplier_id"],
                "supplier_name": hook_out.get("supplier_name"),
                "confidence": float(hook_out.get("confidence", 0.6)),
                "method": "claude",
            }

    return {"supplier_id": None, "supplier_name": None, "confidence": 0.0, "method": "none"}


def load_known_aliases(conn) -> list[dict]:
    """Materialize the alias list a parser needs from a live SQLite connection."""
    rows = conn.execute(
        """
        SELECT s.id AS supplier_id, s.name AS supplier_name,
               s.canonical_name AS normalized, s.canonical_name AS alias
        FROM suppliers s
        UNION ALL
        SELECT s.id AS supplier_id, s.name AS supplier_name,
               a.normalized AS normalized, a.alias AS alias
        FROM supplier_aliases a JOIN suppliers s ON s.id = a.supplier_id
        """
    ).fetchall()
    return [
        {
            "supplier_id": r["supplier_id"],
            "supplier_name": r["supplier_name"],
            "normalized": r["normalized"],
            "alias": r["alias"],
        }
        for r in rows
    ]


# ---- invoice line extraction -------------------------------------------------

def _classify_amount_match(match: re.Match) -> tuple[float, Optional[str]]:
    raw = match.group("num").replace(",", "")
    value = float(raw)
    if match.group("paren"):
        value = -value
    currency = None
    sym = match.group("sym")
    code = match.group("code")
    if sym:
        currency = _CURRENCY_SYMBOL.get(sym)
    elif code:
        currency = code.upper()
    return value, currency


def _looks_like_invoice_number(token: str) -> bool:
    """Heuristic: trim trailing punctuation, then require >=3 chars & a digit."""
    t = token.strip().rstrip(".,;:")
    if len(t) < 3:
        return False
    if not any(ch.isdigit() for ch in t):
        return False
    # Avoid pulling pure dates / amounts back in as invoice tokens.
    if _DATE_RE.fullmatch(t):
        return False
    return True


_HEADER_KEYWORDS = (
    "total", "subtotal", "balance", "amount due", "period",
    "account no", "statement", "vat", "tax",
)


def _extract_from_line(line: str) -> Optional[ParsedRow]:
    """Try to pull an invoice row out of a single text line.

    Requires a date in the line — line-item rows on a supplier statement
    almost always carry one, while summary/header lines (``Statement Total:
    …``, ``Account No: …``) do not. This keeps the false-positive rate low
    without a per-statement template.
    """
    if not line or not line.strip():
        return None
    date_match = _DATE_RE.search(line)
    if not date_match:
        return None

    amounts = list(_AMOUNT_RE.finditer(line))
    # Filter out amounts whose span is fully inside the date match — `2026` etc.
    date_span = date_match.span()
    amounts = [
        a for a in amounts
        if not (a.start() >= date_span[0] and a.end() <= date_span[1])
    ]
    if not amounts:
        return None
    amount_match = amounts[-1]
    amount, currency = _classify_amount_match(amount_match)

    invoice_date = parse_date(date_match.group(1))
    consumed_spans = [amount_match.span(), date_span]

    invoice_number = ""
    for tok_match in _INVOICE_TOKEN_RE.finditer(line):
        span = tok_match.span()
        if any(span[0] < e and span[1] > s for s, e in consumed_spans):
            continue
        # group(0) preserves the prefix (e.g. "INV-") consumed by the optional
        # non-capturing head; group(1) would drop it.
        candidate = tok_match.group(0).strip()
        if _looks_like_invoice_number(candidate):
            invoice_number = candidate.rstrip(".,;:")
            break

    if not invoice_number:
        return None

    return {
        "invoice_number": invoice_number,
        "normalized_number": normalize_invoice_number(invoice_number),
        "invoice_date": invoice_date.isoformat() if invoice_date else None,
        "amount": amount,
        "currency": currency,
        "raw": {"line": line.strip()},
    }


def _row_to_parsed(row: Sequence[str]) -> Optional[ParsedRow]:
    """Try to interpret a table row (list of cells) as an invoice line.

    Strategy: anchor on the amount cell first (the last numeric cell), then
    pick invoice number + date from the remaining cells. Rows whose only
    "amount-like" cell sits where the invoice number should be (e.g. summary
    rows like ``Statement Total | | | 4950.50 | EGP``) yield no other cell
    that qualifies as an invoice number, so they're skipped.
    """
    cells = [str(c).strip() if c is not None else "" for c in row]
    if not any(cells):
        return None

    # Find the amount cell — last cell that parses as a number AND isn't a date.
    amount = None
    currency = None
    amount_idx = -1
    for i in range(len(cells) - 1, -1, -1):
        c = cells[i]
        if not c:
            continue
        if parse_date(c) is not None:
            continue
        try:
            amount = parse_amount(c)
        except (ValueError, TypeError):
            continue
        m = _AMOUNT_RE.search(c)
        if m:
            _, currency = _classify_amount_match(m)
        amount_idx = i
        break
    if amount is None:
        return None

    lowered = " ".join(cells).lower()
    if any(kw in lowered for kw in _HEADER_KEYWORDS):
        return None

    # Currency may live in its own column (e.g. "...,1500.00,EGP").
    if currency is None:
        for i, c in enumerate(cells):
            if i == amount_idx:
                continue
            up = c.strip().upper()
            if up in ("USD", "GBP", "EUR", "EGP"):
                currency = up
                break

    # Invoice number: first remaining cell that looks like one.
    invoice_number = ""
    for i, c in enumerate(cells):
        if i == amount_idx or not c:
            continue
        if parse_date(c) is not None:
            continue
        if _looks_like_invoice_number(c):
            invoice_number = c
            break
    if not invoice_number:
        return None

    # Date: first remaining cell that parses as one.
    invoice_date = None
    for i, c in enumerate(cells):
        if i == amount_idx or not c or c == invoice_number:
            continue
        d = parse_date(c)
        if d is not None:
            invoice_date = d
            break

    return {
        "invoice_number": invoice_number,
        "normalized_number": normalize_invoice_number(invoice_number),
        "invoice_date": invoice_date.isoformat() if invoice_date else None,
        "amount": amount,
        "currency": currency,
        "raw": {"row": cells},
    }


def extract_invoices(
    text: str,
    tables: Optional[list[list[list[str]]]] = None,
    *,
    claude_hook: Optional[ClaudeHook] = None,
) -> list[ParsedRow]:
    """Extract invoice rows from text + optional tables.

    Tables are tried first (they carry the cleanest signal). Then text lines
    are scanned. Duplicates (same normalized number + amount) are coalesced.
    """
    results: list[ParsedRow] = []
    seen: set[tuple[str, float]] = set()

    def push(row: Optional[ParsedRow]) -> None:
        if not row:
            return
        key = (row["normalized_number"], round(float(row["amount"]), 2))
        if key in seen:
            return
        seen.add(key)
        results.append(row)

    table_rows_found = 0
    for tbl in tables or []:
        for row in tbl:
            before = len(results)
            push(_row_to_parsed(row))
            if len(results) > before:
                table_rows_found += 1

    # CSV/Excel/PDF table extraction often includes the same rows again in the
    # raw text. Prefer clean table rows when we found any; scanning both can
    # double-count invoice lines and misread CSV commas/dates as negative amounts.
    if text and table_rows_found == 0:
        for line in text.splitlines():
            push(_extract_from_line(line))

    if not results and claude_hook is not None:
        hook_out = claude_hook("extract_invoices", {"text": text, "tables": tables or []})
        if hook_out and isinstance(hook_out.get("invoices"), list):
            for inv in hook_out["invoices"]:
                number = inv.get("invoice_number", "")
                push({
                    "invoice_number": number,
                    "normalized_number": inv.get("normalized_number") or normalize_invoice_number(number),
                    "invoice_date": inv.get("invoice_date"),
                    "amount": float(inv["amount"]),
                    "currency": inv.get("currency"),
                    "raw": inv.get("raw") or {"source": "claude"},
                })

    return results


# ---- statement-level metadata ------------------------------------------------

def extract_statement_total(text: str) -> dict:
    """Locate ``Statement Total: £18,164.00`` style lines.

    Returns ``{"total", "currency", "period"}`` — any field may be ``None``.
    """
    if not text:
        return {"total": None, "currency": None, "period": None}

    total: Optional[float] = None
    currency: Optional[str] = None
    period: Optional[str] = None

    for line in text.splitlines():
        stripped = line.strip()
        if total is None:
            m = _TOTAL_LINE_RE.match(stripped)
            if m:
                tail = m.group(1)
                amt_match = _AMOUNT_RE.search(tail)
                if amt_match:
                    total, currency = _classify_amount_match(amt_match)
        if period is None:
            pm = _PERIOD_RE.search(stripped)
            if pm:
                period = pm.group(1).strip().rstrip(".,;")
        if total is not None and period is not None:
            break

    return {"total": total, "currency": currency, "period": period}


# ---- top-level orchestration -------------------------------------------------

def parse_statement(
    path: str | Path,
    *,
    known_aliases: Optional[Iterable[dict]] = None,
    claude_hook: Optional[ClaudeHook] = None,
) -> ParsedStatement:
    """One-shot: file -> ParsedStatement.

    The caller is responsible for sourcing ``known_aliases`` (usually via
    :func:`load_known_aliases` on a live DB connection).
    """
    parsed = parse_file(path)
    supplier = identify_supplier(
        parsed.text,
        filename=Path(path).name,
        known_aliases=known_aliases,
        claude_hook=claude_hook,
    )
    invoices = extract_invoices(parsed.text, parsed.tables, claude_hook=claude_hook)
    totals = extract_statement_total(parsed.text)

    # Currency: prefer total-line currency, else most-common per-line currency.
    currency = totals["currency"]
    if not currency:
        counts: dict[str, int] = {}
        for inv in invoices:
            c = inv.get("currency")
            if c:
                counts[c] = counts.get(c, 0) + 1
        if counts:
            currency = max(counts.items(), key=lambda kv: kv[1])[0]

    return ParsedStatement(
        supplier_name_detected=supplier["supplier_name"],
        supplier_id=supplier["supplier_id"],
        supplier_confidence=supplier["confidence"],
        invoices=invoices,
        statement_total=totals["total"],
        statement_period=totals["period"],
        currency=currency,
        invoice_count=len(invoices),
        source=parsed.source,
    )


__all__ = [
    "ClaudeHook",
    "ParsedFile",
    "ParsedStatement",
    "extract_invoices",
    "extract_statement_total",
    "identify_supplier",
    "load_known_aliases",
    "parse_csv",
    "parse_excel",
    "parse_file",
    "parse_pdf",
    "parse_statement",
    "parse_text",
]
